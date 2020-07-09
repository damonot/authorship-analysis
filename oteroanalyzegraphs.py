# -*- coding: utf-8 -*-
"""
Created on Sat Jan 25 18:55:22 2020

@author: damon
"""
import os
import math
import pandas as pd
import openpyxl # 1-indexed, not 0-indexed
from openpyxl import load_workbook
from itertools import permutations

def runner(repo):
    cwd = os.getcwd()
    auth2flawsXL = cwd + '\lynks\oterolynks-'+repo+'-auth2flaws.xlsx'
    auth2filesXL = cwd + '\lynks\oterolynks-'+repo+'-auth2flawedFiles.xlsx'
    coworkersXL = cwd + '\lynks\oterolynks-'+repo+'-coworkers.xlsx'
    
    excelList = [auth2flawsXL, auth2filesXL, coworkersXL]
    
    response = input('Degree Centrality? [y]/n\n')
    if(response == 'y'):
        for xl in excelList:    
            gen_degreeCentrality(xl, repo)
            print('\n')

    bugvulnfolder = os.getcwd() + "\\text_data\\"
    bugvulnfile = 'otero-'+repo+'-auth2flawedFiles.txt'
    bugvulnpath = bugvulnfolder + bugvulnfile    

    # bf*iaf bug uniqueness distribution for each author
    response = input('Bug-Frequency * Inverse Author Frequency? [y]/n\n')
    if(response == 'y'):
        bfiaf = ff_iaf(bugvulnpath, repo, "bug")
        ffiaf2excel(bfiaf, repo, "Bug")

   # vf*iaf vuln uniqueness distribution for each author
    response = input('Vuln-Frequency * Inverse Author Frequency? [y]/n\n')
    if(response == 'y'):
        vfiaf = ff_iaf(bugvulnpath, repo, "vuln")
        ffiaf2excel(vfiaf, repo, "Vulnerability")

    response = input("Bicluster? [y]/n\n")
    if(response == 'y'):
        bicluster(bugvulnpath, repo)

def bicluster(txt, repo):
    print("biclustering")
    data = get_numerical_data(txt, repo)
    print(data)
    network = get_network(data)
    ffiaf2excel(data, repo, "BiCluster")
    network2excel(network, repo, "BiCluster")

def network2excel(network, repo, type):
    folder = "xl_data\\"
def get_network(data):
    allperms = []
    for tup in data:
        flaws = []
        for flaw in tup[1]:
            flaws.append(flaw)
        flawperms = permutations(flaws, 2)
        allperms.append(flawperms)

    allperms = list(dict.fromkeys(allperms)) # remove duplicate entries
    return allperms

def get_numerical_data(txt, repo):
    flaws = []
    authors = []
    flawLines = []
    with open(txt, encoding='utf-8') as t:
        for line in t:
            fields = line.split('\t') # 0auth | 1file | 2flawtype | 3line | 4rule |
            flawLines.append(fields) # copy necessary lines into list
            flaw = fields[4]
            auth = fields[0]
            if(flaw not in flaws): # flaw not encountered yet
                flaws.append(flaw)
            if(auth not in authors): # author not encountered yet
                authors.append(auth)
    
    nAuth = len(authors) # number of authors
    
    #FF calculation
    listDict = [] # list of freqDict
    for auth in authors:
        freqDict = {} # of times each flaw is connected to auth
        for line in flawLines:
            if(line[0] == auth):
                flaw = line[4]
                if(flaw not in freqDict):
                    freqDict[flaw] = 1
                else:
                    freqDict[flaw] +=1
        
        listDict.append([auth, freqDict]) # author & their flaw freq dist
    
    '''structure is...
    list: tups
        tup: author, dict
            dict: flaw, freq '''

    return listDict


def ffiaf2excel(ffiaf, repo, type):
    
    #print(bfiaf)
    folder = "xl_data\\"
    if(type == "Bug"):
        xlname = folder + "otero-bfiaf.xlsx"
    if(type == "Vulnerability"):
        xlname = folder + "otero-vfiaf.xlsx"
    if(type == "AUTHOR"):
        xlname = folder + "otero-"+repo+"-authorProxmeasure.xlsx"
        type = "Bug"
    if(type == "FILE"):
        xlname = folder + "otero-"+repo+"-fileProxmeasure.xlsx"
        type = "Bug"
    if(type == "BiCluster"):
        xlname = folder + "otero-"+repo+"-biconData.xlsx"
        type = "Flaw"
    try:
        book = load_workbook(xlname)
    except:
        book = openpyxl.Workbook()
        default = book.get_sheet_by_name('Sheet')
        book.remove_sheet(default)

    sheet = book.create_sheet(repo)

    # generate column headers
    cl = sheet.cell(row=1, column=1)
    cl.value = type+"_Rule"
    col = 2; row = 1
    for tups in ffiaf:
        cl = sheet.cell(row=row, column=col)
        cl.value = tups[0]
        col+=1

    # generate row headers i.e. rule types
    rules = []
    for tups in ffiaf:
        for rule in tups[1]:
            cleanRule = (rule.replace('\n', ''))
            if(cleanRule not in rules):
                rules.append(cleanRule)
    col = 1; row = 2;
    for rule in rules:
        cl = sheet.cell(row=row, column=col)
        cl.value = rule
        #sheet.write(row,col, rule)
        row+=1

    '''
    structure is...
    list: tups
        tup: author, dict
            dict: flaw, freq 
    '''


    # populate table 
    col = 2
    for tups in ffiaf:
        dict = tups[1]
        for rule in dict:
            row = 2
            cleanRule = (rule.replace('\n', ''))
            cl = sheet.cell(row=row, column=1)
            rowRule = cl.value
            while(rowRule != cleanRule):
                row+=1
                cl = sheet.cell(row=row, column=1)
                rowRule = cl.value
            cl = sheet.cell(row=row, column=col)
            cl.value = dict[rule]
            #print(rule, dict[rule])
        col+=1


    book.save(xlname)


#flaw-frequency2inverseauthorfrequency
def ff_iaf(txt, repo, type):
    flaws = []
    authors = []
    flawLines = []
    with open(txt, encoding='utf-8') as t:
        for line in t:
            fields = line.split('\t') # 0auth | 1file | 2flawtype | 3line | 4rule |
            if(fields[2].strip() == type): # is a bug/vuln? weird formatting, work with it
                flawLines.append(fields) # copy necessary lines into list
                flaw = fields[4]
                auth = fields[0]
                if(flaw not in flaws): # flaw not encountered yet
                    flaws.append(flaw)
                if(auth not in authors): # author not encountered yet
                    authors.append(auth)
    
    nAuth = len(authors) # number of authors
    
    #FF calculation
    listDict = [] # list of freqDict
    for auth in authors:
        freqDict = {} # of times each flaw is connected to auth
        for line in flawLines:
            if(line[0] == auth):
                flaw = line[4]
                if(flaw not in freqDict):
                    freqDict[flaw] = 1
                else:
                    freqDict[flaw] +=1
        
        listDict.append([auth, freqDict]) # author & their flaw freq dist
    
    '''structure is...
    list: tups
        tup: author, dict
            dict: flaw, freq '''
    #AF calculation: num authors who have written a particular flaw
    
    encounteredFlaws = []
    authsPerFlaw = {} # flaw | num auths who have written it
    for tup in listDict:
        auth = tup[0]
        for flaw in tup[1]: # iterate through flaws of auths freqdist dict
            if flaw not in encounteredFlaws: # new author of flaw
                authsPerFlaw[flaw] = 1
                encounteredFlaws.append(flaw)
            else: authsPerFlaw[flaw] += 1 # additional author of existing flaw
    
    
    '''structure is...
    list: tups
        tup: author, dict
            dict: flaw, bfiaf '''
    #FF IAF calculation
    ffiafs = []
    for auth in authors:
        ffiafs.append((auth, {}))
    
    for flaw in authsPerFlaw:
        iafeqn = (1 + nAuth)/(1+authsPerFlaw[flaw])
        iaf = math.log(iafeqn,2)
        for tup in listDict: # (auth, {flaw, count})
            auth = tup[0]
            authIndex = listDict.index(tup)
            if(flaw in tup[1]):
                flawfreq4auth = tup[1][flaw] # [1] selects dict, [flaw] selects key
            else: flawfreq4auth = 0
            ffiafval = flawfreq4auth * iaf
            
            # append bug and bfiafval to dict
            ffiafs[authIndex][1].update( {flaw : ffiafval} ) 
    
    return ffiafs
    

def gen_degreeCentrality(xlLoc, repo):
    print(xlLoc)
    xls = pd.ExcelFile(xlLoc)
    df1 = pd.read_excel(xls, 'relations')
    
    
    auths = {}
    for row in df1['node1']:
        if row not in auths:
            auths[row] = 1
        else:
            auths[row] +=1


    sheetName = xlLoc[xlLoc.rindex('-')+1:] # save text after last dash
    xlName = repo+'-degreecentrality.xlsx'
    df2 = pd.DataFrame(list(auths.items()),columns = ['node','count'])
    df2 = df2.sort_values('count', ascending=False) # sort in descending order
    print('output xl: '+xlName)
    df2excel(df2, xlName, sheetName) # convert dataframe to excel doc


# convert dataframe to excel
def df2excel(df, xlName, sheetName):
    
    if not(os.path.isfile(xlName)):
        print(xlName+' DNE')
        writer = pd.ExcelWriter(xlName, engine='xlsxwriter')
        writer.save()
    else: print() # do nothing, xlName already exists
        
    with pd.ExcelWriter(xlName, engine='openpyxl') as writer:
        writer.book = load_workbook(xlName)
        df.to_excel(writer, sheetName, index=False)
    


    
#oterolynks-phpmyadmin-auth2flaws.xlsx


