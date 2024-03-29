# -*- coding: utf-8 -*-
"""
Created on Sat Jan 25 18:55:22 2020

@author: damon
"""
import os
import math
import pandas as pd
import openpyxl # 1-indexed, not 0-indexed
from openpyxl import load_workbook
from itertools import permutations
import scripts.oteromakegraphdata as mkgrf


def auth_influence(verbose, overwrite, repo):
    authinftxt = os.getcwd() + '\output\{}\{}-authinfluence.txt'.format(repo,repo)
    if not overwrite:
        if(os.path.isfile(authinftxt)):
                response = input("Overwrite {}? File already exists. Respond [y]/n".format(authinftxt))
        else: response = 'y'
    else:
        response = 'y'

    if response == 'y':
        if verbose:
            print('calculating author influence for {}'.format(repo))

        # auth influence = [ authflaw / total flaws ] * [auth flawed files / total flawed files ]
        
        authflawtxt = os.getcwd() + '\output\{}\{}-authflaw.txt'.format(repo, repo)

        if not(os.path.isfile(authflawtxt)):
            mkgrf.auth_flaw(verbose, overwrite, repo)

        authflawdict, totalFlaws = flaws_per_auth(verbose, overwrite, repo, authflawtxt)  
        authfiledict, totalFiles = files_per_auth(verbose, overwrite, repo, authflawtxt)

        # update values 
        for auth in authflawdict:
            authflaws = authflawdict[auth]
            authflawdict[auth] = authflaws / totalFlaws

        for auth in authfiledict:
            authfiles = authfiledict[auth]
            authfiledict[auth] = authfiles / totalFiles

        # multiply values together
        authinfluence = {}
        for auth, count in authflawdict.items():
            infVal = round(count * authfiledict[auth], 5)
            authinfluence[auth] = infVal

        dict_to_txt(verbose, overwrite, repo, authinfluence, authinftxt)


def dict_to_txt(verbose, overwrite, repo, dict, txt):
    if verbose:
        print("writing {}".format(txt))

    with open(txt, "w", encoding='utf-8') as output:
        for key,val in dict.items():
                output.write('{}\t{}\n'.format(key, val))


def flaws_per_auth(verbose, overwrite, repo, fyle):
    if verbose:
        print("\tcounting flaws per author for {}".format(repo))

    authflawdict = {}
    totalFlawsCount = 0
    with open (fyle, encoding='utf-8') as f:
        for line in f:
            fields = line.split('\t') # 0auth | 1file | 2flawtype | 3line | 4rule |
            totalFlawsCount+=1
            auth = fields[0].rstrip()
            
            if auth not in authflawdict:
                authflawdict[auth] = 1
            else:
                authflawdict[auth] +=1

    return authflawdict, totalFlawsCount


def files_per_auth(verbose, overwrite, repo, fyle):
    if verbose:
        print("\tcounting files per author for {}".format(repo))

    files = [] # len(files) = total
    authfiletups = []
    authfiledict = {}

    with open (fyle, encoding='utf-8') as f:
        for line in f:
            fields = line.split('\t') # 0auth | 1file | 2flawtype | 3line | 4rule |
            auth = fields[0].rstrip()
            flawedFile = fields[1]
            if flawedFile not in files:
                files.append(flawedFile)

            tup = (auth, flawedFile)
            if tup not in authfiletups:
                authfiletups.append(tup)
                if auth not in authfiledict:
                    authfiledict[auth] = 1
                else:
                    authfiledict[auth] +=1
                

    return authfiledict, len(files)


def centrality(verbose, overwrite, repo):

    authflawtxt = os.getcwd() + '\output\{}\{}-authflaw.txt'.format(repo, repo)

    # authfile centrality
    authfilecentxt = os.getcwd() + '\output\{}\{}-authfilecentraltiy.txt'.format(repo,repo)
    if not overwrite:
        if(os.path.isfile(authfilecentxt)):
                response = input("Overwrite {}? File already exists. Respond [y]/n".format(authfilecentxt))
        else: response = 'y'
    else:
        response = 'y'

    if response == 'y':
        if verbose:
            print("calculating centrality for {} according to file count".format(repo))
        filesdict, totalFiles = files_per_auth(verbose, overwrite, repo, authflawtxt)

        # update values 
        for auth in filesdict:
            authfiles = filesdict[auth]
            filesdict[auth] = round(authfiles / totalFiles, 5)

        dict_to_txt(verbose, overwrite, repo, filesdict, authfilecentxt)

    # authflaw centrality
    authflawcentxt = os.getcwd() + '\output\{}\{}-authflawcentraltiy.txt'.format(repo,repo)
    if not overwrite:
        if(os.path.isfile(authflawcentxt)):
                response = input("Overwrite {}? File already exists. Respond [y]/n".format(authflawcentxt))
        else: response = 'y'
    else:
        response = 'y'

    if response == 'y':
        if verbose:
            print("calculating centrality for {} according to flaw count".format(repo))
        flawsdict, totalFlaws = flaws_per_auth(verbose, overwrite, repo, authflawtxt) 

        # update values 
        for auth in flawsdict:
            authflaws = flawsdict[auth]
            flawsdict[auth] = round(authflaws / totalFlaws, 5)

        dict_to_txt(verbose, overwrite, repo, flawsdict, authflawcentxt)


def ffiaf(verbose, overwrite, repo):
    types = ["flaw", "bug", "vuln"] 
    for type in types:
        ffiaf_permutations(verbose, overwrite, repo, type)


def ffiaf_permutations(verbose, overwrite, repo, type):
    ffiafXL = os.getcwd() + '\output\{}\{}-{}-fiaf.xlsx'.format(repo,repo, type)
    
    if not overwrite:
        if(os.path.isfile(ffiafXL)):
                response = input("Overwrite {}? File already exists. Respond [y]/n".format(ffiafXL))
        else: response = 'y'
    else:
        response = 'y'

    if response == 'y':
        if verbose:
            print("{}-fiaf for {}".format(type, repo))

        authflaw = os.getcwd() + '\output\{}\{}-authflaw.txt'.format(repo, repo)
        ffiaf = calc_ffiaf(authflaw, repo, type)
        ffiaf2excel(ffiaf, repo, type)


#flaw-frequency2inverseauthorfrequency
def calc_ffiaf(txt, repo, type):
    flaws = []
    authors = []
    flawLines = []
    with open(txt, encoding='utf-8') as t:
        for line in t:
            fields = line.split('\t') # 0auth | 1file | 2flawtype | 3line | 4rule |
            if(fields[2].strip() == type) or (type == 'flaw'): # is a bug/vuln? weird formatting, work with it
                flawLines.append(fields) # copy necessary lines into list
                
                if fields[2].strip() == 'bug':
                    prefix = 'B-'
                elif fields[2].strip() == 'vuln':
                    prefix = 'V-'

                flaw = prefix + fields[4]
                auth = fields[0]
                if(flaw not in flaws): # flaw not encountered yet
                    flaws.append(flaw)
                if(auth not in authors): # author not encountered yet
                    authors.append(auth)
    
    nAuth = len(authors) # number of authors
    
    #FF calculation
    listDict = [] # list of freqDict
    for auth in authors:
        freqDict = {} # of times each flaw is connected to auth
        for line in flawLines:
            if(line[0] == auth):

                if line[2].strip() == 'bug':
                    prefix = 'B-'
                elif line[2].strip() == 'vuln':
                    prefix = 'V-'

                flaw = prefix + line[4]
                if(flaw not in freqDict):
                    freqDict[flaw] = 1
                    #print(flaw)
                else:
                    freqDict[flaw] +=1
        
        listDict.append([auth, freqDict]) # author & their flaw freq dist
    
    '''structure is...
    list: tups
        tup: author, dict
            dict: flaw, freq '''
    #AF calculation: num authors who have written a particular flaw
    
    encounteredFlaws = []
    authsPerFlaw = {} # flaw | num auths who have written it
    for tup in listDict:
        auth = tup[0]
        for flaw in tup[1]: # iterate through flaws of auths freqdist dict
            if flaw not in encounteredFlaws: # new author of flaw
                authsPerFlaw[flaw] = 1
                encounteredFlaws.append(flaw)
            else: authsPerFlaw[flaw] += 1 # additional author of existing flaw
    
    
    '''structure is...
    list: tups
        tup: author, dict
            dict: flaw, bfiaf '''
    #FF IAF calculation
    ffiafs = []
    for auth in authors:
        ffiafs.append((auth, {}))
    
    for flaw in authsPerFlaw:
        iafeqn = (1 + nAuth)/(1+authsPerFlaw[flaw])
        iaf = math.log(iafeqn,2)
        for tup in listDict: # (auth, {flaw, count})
            auth = tup[0]
            authIndex = listDict.index(tup)
            if(flaw in tup[1]):
                flawfreq4auth = tup[1][flaw] # [1] selects dict, [flaw] selects key
            else: flawfreq4auth = 0
            ffiafval = flawfreq4auth * iaf
            
            # append bug and bfiafval to dict
            ffiafs[authIndex][1].update( {flaw : ffiafval} ) 
    
    return ffiafs


def ffiaf2excel(ffiaf, repo, type):
    

    folder = 'xl_data\\'
    if(type == "bug"):
        xlname = os.getcwd() + '\output\{}\{}-{}-fiaf.xlsx'.format(repo,repo, type)
    if(type == "vuln"):
        xlname = os.getcwd() + '\output\{}\{}-{}-fiaf.xlsx'.format(repo,repo, type)
    if(type == "flaw"):
        xlname = os.getcwd() + '\output\{}\{}-{}-fiaf.xlsx'.format(repo,repo, type)
    if(type == "AUTHOR"):
        xlname = folder + "otero-"+repo+"-authorProxmeasure.xlsx"
        type = "bug"
    if(type == "FILE"):
        xlname = folder + "otero-"+repo+"-fileProxmeasure.xlsx"
        type = "bug"
    if(type == "BiCluster"):
        xlname = folder + "otero-"+repo+"-biconData.xlsx"
        type = "flaw"
    try:
        book = load_workbook(xlname)
    except:
        book = openpyxl.Workbook()
        default = book.get_sheet_by_name('Sheet')
        book.remove_sheet(default)

    sheet = book.create_sheet(repo)

    # generate column headers
    cl = sheet.cell(row=1, column=1)
    cl.value = type+"_rule"
    col = 2; row = 1
    for tups in ffiaf:
        cl = sheet.cell(row=row, column=col)
        cl.value = tups[0]
        col+=1

    # generate row headers i.e. rule types
    rules = []
    for tups in ffiaf:
        for rule in tups[1]:
            cleanRule = (rule.replace('\n', ''))
            if(cleanRule not in rules):
                rules.append(cleanRule)
    col = 1; row = 2;
    for rule in rules:
        cl = sheet.cell(row=row, column=col)
        cl.value = rule
        #sheet.write(row,col, rule)
        row+=1

    '''
    structure is...
    list: tups
        tup: author, dict
            dict: flaw, freq 
    '''


    # populate table 
    col = 2
    for tups in ffiaf:
        dict = tups[1]
        for rule in dict:
            row = 2
            cleanRule = (rule.replace('\n', ''))
            cl = sheet.cell(row=row, column=1)
            rowRule = cl.value
            while(rowRule != cleanRule):
                row+=1
                cl = sheet.cell(row=row, column=1)
                rowRule = cl.value
            cl = sheet.cell(row=row, column=col)
            cl.value = dict[rule]
            #print(rule, dict[rule])
        col+=1


    book.save(xlname)


def runner(repo):
    cwd = os.getcwd()
    auth2flawsXL = cwd + '\lynks\oterolynks-'+repo+'-auth2flaws.xlsx'
    auth2filesXL = cwd + '\lynks\oterolynks-'+repo+'-auth2flawedFiles.xlsx'
    coworkersXL = cwd + '\lynks\oterolynks-'+repo+'-coworkers.xlsx'
    
    excelList = [auth2flawsXL, auth2filesXL, coworkersXL]
    
    response = input('Degree Centrality? [y]/n\n')
    if(response == 'y'):
        for xl in excelList:    
            gen_degreeCentrality(xl, repo)
            print('\n')

    bugvulnfolder = os.getcwd() + "\\text_data\\"
    bugvulnfile = 'otero-'+repo+'-auth2flawedFiles.txt'
    bugvulnpath = bugvulnfolder + bugvulnfile    

    # bf*iaf bug uniqueness distribution for each author
    response = input('Bug-Frequency * Inverse Author Frequency? [y]/n\n')
    if(response == 'y'):
        bfiaf = ff_iaf(bugvulnpath, repo, "bug")
        ffiaf2excel(bfiaf, repo, "Bug")

   # vf*iaf vuln uniqueness distribution for each author
    response = input('Vuln-Frequency * Inverse Author Frequency? [y]/n\n')
    if(response == 'y'):
        vfiaf = ff_iaf(bugvulnpath, repo, "vuln")
        ffiaf2excel(vfiaf, repo, "Vulnerability")

def network2excel(network, repo, type):
    folder = "xl_data\\"

    xlname = folder + "otero-"+repo+"-biconExprs.xlsx"
    try:
        book = load_workbook(xlname)
    except:
        book = openpyxl.Workbook()
        default = book.get_sheet_by_name('Sheet')
        book.remove_sheet(default)

    sheet = book.create_sheet(repo)


    row = 1
    for conn in network:
        col = 1
        cl = sheet.cell(row=row, column=col)
        cl.value = conn[0]
        col+=1
        cl = sheet.cell(row=row, column=col)
        cl.value = conn[1]
        row+=1

    book.save(xlname)


def get_network(data):
    allperms = []
    for tup in data:
        flaws = []
        for flaw in tup[1]:
            flaws.append(flaw.rstrip())
        flawperms = permutations(flaws, 2)
        for perm in list(flawperms):
            if(perm not in allperms):
                allperms.append(perm)
    
    print(allperms)
    return allperms

def get_numerical_data(txt, repo):
    flaws = []
    authors = []
    flawLines = []
    with open(txt, encoding='utf-8') as t:
        for line in t:
            fields = line.split('\t') # 0auth | 1file | 2flawtype | 3line | 4rule |
            flawLines.append(fields) # copy necessary lines into list
            flaw = fields[4]
            auth = fields[0]
            if(flaw not in flaws): # flaw not encountered yet
                flaws.append(flaw)
            if(auth not in authors): # author not encountered yet
                authors.append(auth)
    
    nAuth = len(authors) # number of authors
    
    #FF calculation
    listDict = [] # list of freqDict
    for auth in authors:
        freqDict = {} # of times each flaw is connected to auth
        for line in flawLines:
            if(line[0] == auth):
                flaw = line[4]
                if(flaw not in freqDict):
                    freqDict[flaw] = 1
                else:
                    freqDict[flaw] +=1
        
        listDict.append([auth, freqDict]) # author & their flaw freq dist
    
    '''structure is...
    list: tups
        tup: author, dict
            dict: flaw, freq '''

    return listDict


def gen_degreeCentrality(xlLoc, repo):
    print(xlLoc)
    xls = pd.ExcelFile(xlLoc)
    df1 = pd.read_excel(xls, 'relations')
    
    
    auths = {}
    for row in df1['node1']:
        if row not in auths:
            auths[row] = 1
        else:
            auths[row] +=1


    sheetName = xlLoc[xlLoc.rindex('-')+1:] # save text after last dash
    xlName = repo+'-degreecentrality.xlsx'
    df2 = pd.DataFrame(list(auths.items()),columns = ['node','count'])
    df2 = df2.sort_values('count', ascending=False) # sort in descending order
    print('output xl: '+xlName)
    df2excel(df2, xlName, sheetName) # convert dataframe to excel doc


# convert dataframe to excel
def df2excel(df, xlName, sheetName):
    
    if not(os.path.isfile(xlName)):
        print(xlName+' DNE')
        writer = pd.ExcelWriter(xlName, engine='xlsxwriter')
        writer.save()
    else: print() # do nothing, xlName already exists
        
    with pd.ExcelWriter(xlName, engine='openpyxl') as writer:
        writer.book = load_workbook(xlName)
        df.to_excel(writer, sheetName, index=False)
    

