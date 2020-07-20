# -*- coding: utf-8 -*-
"""
Created on Sat Jan  4 00:21:57 2020

@author: damon
"""

import pandas as pd
import xlsxwriter
from openpyxl import load_workbook
import os


def help():
    print("'oteromakelynsoft' generates an .xlsx formatted for Lynksoft")
    
    response = input("To see neccesary fields, press 1. To see all fields press 2.")
    if(response == '2'):
        print("person: label, color, scale,	tags, extras\n"+
              "relation types: description, color, style, thickness\n"+
              "relations: node1, node2, edge_type")    
    elif(response == '2'):
        print("document (optional): label, color, scale, extras\n"+
              "flag (optional): label, color, scale, extras\n"+
              "group: label, color, scale, extras\n"+
              "location (optional): label, color, scale, extras\n"+
              "organization (optional): label, color, scale, extras\n"+
              "person: label, color, scale,	tags, extras\n"+
              "project (optional): label, color, scale, extras\n"+
              "relation types: description, color, style, thickness\n"+
              "relations: node1, node2, edge_type\n"+
              "time (optional) label, color, scale, extras")

def gen_generic_sheet():
    return pd.DataFrame(columns=['label', 'color', 'scale', 'tags', 'extras'])

# generate dataframe of authors
def gen_personDF(personList):
    personDF = gen_generic_sheet() # generic sheet
    for person in personList: # add email data
        personDF = personDF.append({'label': person, 'color': '#333333', 'scale':'medium'}, ignore_index=True)    
    return personDF
    

# generate dataframe of files containing vulns/bugs/both
def gen_documentDF(documentList):
    documentDF = gen_generic_sheet() # generic sheet
    
    for doc in documentList: # add email data
        color = '#333333' # black
        if(doc[1] == 'vuln' or 'vuln2vuln' in doc[1]):
            color = '#E01A1C' # red
        elif(doc[1] == 'both'):
            color = '#6A39DA' # purple
        else: #(doc[1] == 'bug' or 'vuln2' in doc[1]):
            color = '#00A2FF' # blue
        documentDF = documentDF.append({'label': doc[0], 'color': color, 'scale':'medium'}, ignore_index=True)    
    return documentDF


def gen_flagDF(flagList):
    #input format: 0file, 1type, 2line, 3description
    flagDF = gen_generic_sheet()
    
    #TODO fix lynsoft timeout when tags/extras are included
    
    color = '#333333'
    for flag in flagList:
        label = flag[2] + ' @ ' + flag[0]
        tags = ''#flag[1] #+ ', ' + flag[3]
        message = flag[3]
        extras = ''#'{"details": "' + message + '\n"}'
        if(flag[1] == 'vuln'):
            color = '#E01A1C' # red
        elif(flag[1] == 'bug'):
            color = '#00A2FF' # blue
        flagDF = flagDF.append({'label': label, 'color': color, 'scale':'medium', 'tags': tags, 'extras': extras}, ignore_index=True)    
    return flagDF

        
def gen_relTypeDF(relationTypesList):
    
    color = '#000000'
    style = 'solid'
    thickness = 'small'
    
    
    relTypeDF = pd.DataFrame(columns=['description', 'color', 'style', 'thickness'])
    for relType in relationTypesList:
        if(relType == 'vuln'):
            color = '#333333'
            style = 'solid'
            thickness = 'medium'
        elif(relType == 'bug'):
            color = '#9B9B9B'
            style = 'dashed' 
            thickness = 'small'
        elif(relType == 'common_contributor'): 
                color = '#000000'
                style = 'dashed'
                thickness = 'small'
        relTypeDF = relTypeDF.append({'description': relType, 'color': color, 'style': style, 'thickness':thickness}, ignore_index=True)
    return relTypeDF

def merge_documents(documentList):
    
    duplicateList = []
    
    # find duplicates
    for doc in documentList:
        docName = doc[0]
        docType = doc[1]
        for comp in documentList[:]:
            if(comp[0] == docName and comp[1] != docType):
                doc[1] = 'both'
                duplicateList.append(docName)
    
    newList = []
    # remove duplicates
    for doc in documentList[:]:
        if(doc not in newList):
            newList.append(doc)
            
    return newList
    
    # note: what kind of relationships does the document hold?
def genXLSX(file, columnType, xlsxName):    
    workbook = xlsxwriter.Workbook(xlsxName)
    workbook.close()
    
    personList = [] # list of left labels
    documentList = [] # list of right labels + label type
    flagList = [] # list of vulnerable code snippets (right label)
    dummyList = [] # simplified version of flagList w/ only file and line #
    relationsList = [] # list containing node1, node2, relationshipType
    relationTypesList = []
    #fileIndex = 0
    
    # extract fields from relationship documents
    with open(file, encoding='utf-8') as f:
        for line in f:
            fields = line.split()
            if(len(fields) < 2): # check/skip improper formatting 
                continue
            if(fields[2] not in relationTypesList):
                relationTypesList.append(fields[2])
            if(columnType[0] == 'person'): # left column is person?
                if(fields[0] not in personList): # prevent duplicate persons
                    personList.append(fields[0]) # person/email is the first field
            if(columnType[0] == 'document'): # left column is file?
                if([fields[0], fields[2]] not in documentList): # prevent duplicate. 
                    documentList.append([fields[0], fields[2]]) # project/erroneous file
            if(columnType[1] == 'person'): # right column is person?
                if(fields[1] not in personList): # prevent duplicate persons
                    personList.append(fields[1]) # person/email is the first field
            if(columnType[1] == "document"): # right column is document?
                if([fields[0], fields[2]] not in documentList): # prevent duplicate.
                    documentList.append([fields[1], fields[2]]) # project/erroneous file
            if(columnType[1] == "flag"): # right column is code flaw?
                if([fields[1], fields[3]] not in dummyList): # prevent duplicate. 
                    dummyList.append([fields[1], fields[3]]) # dirty workaround. doesn't track multiple errors on single line
                    flagList.append([fields[1], fields[2], fields[3], fields[4]]) # specific flaw
            
            #format auth, flaw, type, line, desc
            relationship = []
            revrelationship = [] # if this breaks, maybe switch the R back to lowercase?
            if(columnType[1] == "flag"):
                flaw = (str(fields[3])+' @ '+str(fields[1]))
                relationship = [fields[0], flaw, fields[2]]
                revRelationship = [flaw, fields[0], fields[2]]
            else:
                relationship = [fields[0], fields[1], fields[2]]
                revRelationship = [fields[1], fields[0], fields[2]]
            if(relationship not in relationsList and revRelationship not in relationsList):
                relationsList.append(relationship)
    
    # resolve duplicates
    documentList = merge_documents(documentList)


    # generate input sheets
    personDF = gen_personDF(personList) # persons/emails
    documentDF = gen_documentDF(documentList) # files containing vulns/bugs
    flagDF = gen_flagDF(flagList) #TODO fill out flagDF (flaw DF)
    relationsDF = pd.DataFrame(relationsList, columns=['node1','node2','edge_type']) # relationships
    relatTypeDF = gen_relTypeDF(relationTypesList) # relationship types
    
    excelPath = os.getcwd() + '\\' + xlsxName
    

    
    # re-open excel book w/o erasing previous data
    book = load_workbook(excelPath)
    writer = pd.ExcelWriter(excelPath, engine = 'openpyxl')
    writer.book = book

    #bold = writer.add_format({'bold': False})
    
    personDF.to_excel(writer, sheet_name = 'person',  index=False)
    documentDF.to_excel(writer, sheet_name = 'document',  index=False)
    flagDF.to_excel(writer, sheet_name = 'flag',  index=False)
    relatTypeDF.to_excel(writer, sheet_name = 'relation types',  index=False)
    relationsDF.to_excel(writer, sheet_name = 'relations',  index=False)
    
    writer.save()
    writer.close() 
    print('Done. Please unbolden headers before uploading to Lynksoft.')
    
    