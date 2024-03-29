import os
import sys
from bicon import data_preprocessing
from bicon import BiCoN
from bicon import results_analysis
import scripts.check as check
from itertools import permutations
from openpyxl import load_workbook
import openpyxl # 1-indexed, not 0-indexed

def main(verbose, overwrite, repo):
    if verbose:
        print("bicluster main -v {} -o {} repo {}".format(verbose, overwrite, repo))

    gather_data(verbose, overwrite, repo)


def gather_data(verbose, overwrite, repo):
    if verbose:
        print("gathering files for {}".format(repo))
    cwd = os.getcwd()
    
    txt = '{}\output\{}\{}-authflaw.txt'.format(cwd, repo, repo)
    print(txt)
    expdata = get_numerical_data(txt, repo)

    exists = check.fyle(verbose, repo, txt)
    if not exists:
        generate_dependency(verbose, repo)


    #func = getattr(mkgrf, 'ffiaf')
    #func(verbose, overwrite, repo)

    fol = "{}\\bicluster\\".format(repo)
    print(fol)

    check.folder(verbose, fol)

    network = get_network(expdata)
    expdata2excel(expdata, repo, "BiCluster")
    network2excel(network, repo, "BiCluster")

    bicon_analysis(repo)


def generate_dependency(verbose, repo):
    if verbose:
        print("'{}-authflaw.txt' not found. Generating now...".format(repo))
    
    #TODO call (venv) python authanalys.py -af -o through cmd

# degree Based
def bicon_analysis(repo):
    print("biclustering for {}".format(repo))

    folder = os.getcwd() + '\output\{}\\bicluster\\'.format(repo)

    #path_expr = folder + '{}-biconExprs.csv'.format(repo)
    path_expr = folder + '{}-flaw-fiaf.csv'.format(repo)
    path_net = folder + '{}-biconNetwork.tsv'.format(repo)


    GE,G,labels, _= data_preprocessing(path_expr, path_net)
    L_g_min = 3
    L_g_max = 15
    model = BiCoN(GE,G,L_g_min,L_g_max)
    solution,scores= model.run_search()

    results = results_analysis(solution, labels)
    results.convergence_plot(scores)

    resultLoc = folder + repo+"-biconResults.csv"
    results.save(output = resultLoc)

    netOut = folder + repo+"-biconNetwork.png"
    results.show_networks(GE, G, output = netOut)
    
    clusterOut = folder + repo+"-biconClustermap.png"
    results.show_clustermap(GE, G, output = clusterOut)


def get_numerical_data(txt, repo):
    flaws = []
    authors = []
    flawLines = []
    with open(txt, encoding='utf-8') as t:
        for line in t:
            fields = line.split('\t') # 0auth | 1file | 2flawtype | 3line | 4rule |
            flawLines.append(fields) # copy necessary lines into list
            flaw = fields[4]
            auth = fields[0]
            if(flaw not in flaws): # flaw not encountered yet
                flaws.append(flaw)
            if(auth not in authors): # author not encountered yet
                authors.append(auth)
    
    nAuth = len(authors) # number of authors
    
    #FF calculation
    listDict = [] # list of freqDict
    for auth in authors:
        freqDict = {} # of times each flaw is connected to auth
        for line in flawLines:
            if(line[0] == auth):
                flaw = line[4]
                if(flaw not in freqDict):
                    freqDict[flaw] = 1
                else:
                    freqDict[flaw] +=1
        
        listDict.append([auth, freqDict]) # author & their flaw freq dist
    
    '''structure is...
    list: tups
        tup: author, dict
            dict: flaw, freq '''

    return listDict


def get_network(data):
    allperms = []
    for tup in data:
        flaws = []
        for flaw in tup[1]:
            flaws.append(flaw.rstrip())
        flawperms = permutations(flaws, 2)
        for perm in list(flawperms):
            if(perm not in allperms):
                allperms.append(perm)
    
    print(allperms)
    return allperms


def expdata2excel(ffiaf, repo, type):
    
    #print(bfiaf)
    folder = '{}\output\{}\\bicluster\\'.format(os.getcwd(), repo)
    if(type == "Bug"):
        xlname = folder + "otero-bfiaf.xlsx"
    if(type == "Vulnerability"):
        xlname = folder + "otero-vfiaf.xlsx"
    if(type == "AUTHOR"):
        xlname = folder + "otero-"+repo+"-authorProxmeasure.xlsx"
        type = "Bug"
    if(type == "FILE"):
        xlname = folder + "otero-"+repo+"-fileProxmeasure.xlsx"
        type = "Bug"
    if(type == "BiCluster"):
        xlname = os.getcwd() + "\output\{}\\bicluster\{}-biconExprs.xlsx".format(repo, repo)
        type = "Flaw"
    try:
        book = load_workbook(xlname)
    except:
        book = openpyxl.Workbook()
        default = book.get_sheet_by_name('Sheet')
        book.remove_sheet(default)

    sheet = book.create_sheet(repo)

    # generate column headers
    cl = sheet.cell(row=1, column=1)
    cl.value = type+"_Rule"
    col = 2; row = 1
    for tups in ffiaf:
        cl = sheet.cell(row=row, column=col)
        cl.value = tups[0]
        col+=1

    # generate row headers i.e. rule types
    rules = []
    for tups in ffiaf:
        for rule in tups[1]:
            cleanRule = (rule.replace('\n', ''))
            if(cleanRule not in rules):
                rules.append(cleanRule)
    col = 1; row = 2;
    for rule in rules:
        cl = sheet.cell(row=row, column=col)
        cl.value = rule
        #sheet.write(row,col, rule)
        row+=1

    '''
    structure is...
    list: tups
        tup: author, dict
            dict: flaw, freq 
    '''


    # populate table 
    col = 2
    for tups in ffiaf:
        dict = tups[1]
        for rule in dict:
            row = 2
            cleanRule = (rule.replace('\n', ''))
            cl = sheet.cell(row=row, column=1)
            rowRule = cl.value
            while(rowRule != cleanRule):
                row+=1
                cl = sheet.cell(row=row, column=1)
                rowRule = cl.value
            cl = sheet.cell(row=row, column=col)
            cl.value = dict[rule]
            #print(rule, dict[rule])
        col+=1


    book.save(xlname)


def network2excel(network, repo, type):
    folder = '{}\output\{}\\bicluster\\'.format(os.getcwd(), repo)

    #xlname = folder + "otero-"+repo+"-biconExprs.xlsx"
    xlname = os.getcwd() + "\output\{}\\bicluster\{}-biconNetwork.xlsx".format(repo, repo)
    try:
        book = load_workbook(xlname)
    except:
        book = openpyxl.Workbook()
        default = book.get_sheet_by_name('Sheet')
        book.remove_sheet(default)

    sheet = book.create_sheet(repo)


    row = 1
    for conn in network:
        col = 1
        cl = sheet.cell(row=row, column=col)
        cl.value = conn[0]
        col+=1
        cl = sheet.cell(row=row, column=col)
        cl.value = conn[1]
        row+=1

    book.save(xlname)


if __name__ == '__main__':
    main(sys.argv[1], sys.argv[2], sys.argv[3])
